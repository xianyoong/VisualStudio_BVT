#!/usr/bin/env python3
"""
Convert an Excel test-case workbook into the CSV the generator consumes.

Usage:
    python tools/xlsx_to_csv.py [path-to-xlsx] [path-to-csv]

Defaults:
    in  = VisualStudio_TC.xlsx        (first worksheet)
    out = VisualStudio_TC_standard.csv

This lets you author test cases directly in Excel (same 17-column layout as the CSV):
just save the workbook and run this script — no manual "Save As CSV" needed.

Requires: openpyxl  (pip install openpyxl)
"""
import csv
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required. Install it with:  python -m pip install openpyxl")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_XLSX = os.path.join(ROOT, "VisualStudio_TC.xlsx")
DEFAULT_CSV = os.path.join(ROOT, "VisualStudio_TC_standard.csv")

# The pipeline schema is fixed at 17 columns (A..Q).
COLUMNS = 17


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def convert(xlsx_path: str, csv_path: str) -> int:
    if not os.path.exists(xlsx_path):
        sys.exit(f"Excel file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active  # first worksheet

    rows_out = 0
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            cells = [cell_text(c) for c in row]
            # Drop fully empty trailing rows; pad/truncate to the fixed 17 columns.
            if not any(cells):
                continue
            cells = (cells + [""] * COLUMNS)[:COLUMNS]
            writer.writerow(cells)
            rows_out += 1

    print(f"Converted {xlsx_path}\n      ->  {csv_path}  ({rows_out} row(s))")
    return rows_out


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    csv_out = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_CSV
    convert(xlsx, csv_out)
